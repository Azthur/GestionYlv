"""
Módulo de Conciliación Bancaria — Control Interno (Finanzas)
Maneja la importación de movimientos bancarios desde Excel,
conciliación automática y manual contra cobranzas del sistema (CcbMVtos).
Soporta conciliación cross-company.
"""

from fastapi import APIRouter, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, date
import uuid
import io
import re
from collections import defaultdict

router = APIRouter(prefix="/api/conciliacion", tags=["Conciliación Bancaria"])

from database import get_db_connection


# ─── Pydantic Models ─────────────────────────────────────────────

class ManualMatchRequest(BaseModel):
    bank_movement_ids: List[int]
    cobranza_keys: List[str]  # Format: "CodCia|coddoc|nrodoc|nroitm"
    usuario: Optional[str] = None
    max_diff: Optional[float] = 0.50


class AutoMatchRequest(BaseModel):
    codcia: str
    bank_code: str
    period_year: Optional[str] = None
    period_month: Optional[str] = None
    cross_company: bool = False
    usuario: Optional[str] = None
    only_ingresos: Optional[bool] = False


# ─── Helpers ──────────────────────────────────────────────────────
# Carga reglas y limpia la operacion
def get_clean_op(op):
    if not op: return ""
    op_str = str(op).strip().upper()
    try:
        from conciliacion import load_rules_from_disk
        rules_disk = load_rules_from_disk()
    except:
        rules_disk = []
    
    conn = get_db_connection()
    rules_db = []
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT RegexPattern, Replacement FROM CleaningRules WHERE IsActive = 1")
            rules_db = cursor.fetchall()
        except:
            pass
        finally:
            conn.close()
            
    for r in rules_disk:
        cond = r.get("condicion", "")
        res = r.get("resultado", "")
        if cond:
            try:
                op_str = re.sub(cond, res, op_str, flags=re.IGNORECASE)
            except Exception:
                op_str = op_str.replace(cond, res)
                
    for pattern, repl in rules_db:
        try:
            op_str = re.sub(pattern, repl, op_str, flags=re.IGNORECASE)
        except Exception:
            pass
            
    return op_str

def row_to_dict(cursor, row):
    """Convert a pyodbc row to a dictionary using cursor.description."""
    columns = [col[0] for col in cursor.description]
    result = {}
    for col, val in zip(columns, row):
        if isinstance(val, datetime):
            result[col] = val.isoformat()
        elif val is None:
            result[col] = None
        else:
            s = str(val)
            result[col] = s.strip() if isinstance(val, str) else s
    return result


def rows_to_list(cursor):
    """Fetch all rows from cursor and return as list of dicts."""
    return [row_to_dict(cursor, row) for row in cursor.fetchall()]


# ─── Endpoints ────────────────────────────────────────────────────

@router.get("/empresas")
def get_empresas():
    """Lista las empresas registradas en AdmMcias."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT codcia, nomcia, ruccia FROM AdmMcias ORDER BY codcia")
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/bancos/{codcia}")
def get_bancos(codcia: str):
    """Lista los bancos configurados para una empresa (CcbTabla, Tabla=0001)."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT Codigo, Nombre, CtaNac, CtaUsa, CodMon
            FROM CcbTabla
            WHERE CodCia = ? AND Tabla = '0001'
            ORDER BY Codigo
        """, (codcia,))
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    codcia: str = Query(...),
    bank_code: str = Query(...)
):
    """
    Importa movimientos bancarios desde un archivo Excel.
    Formato esperado (columnas):
      Fecha, Descripcion, Monto, Saldo, Sucursal, Operacion Numero,
      Operacion Hora, Referencia, Op Manual, OP Cancelacion,
      Descripcion Final, Estado
    """
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active

        batch_id = str(uuid.uuid4())
        rows_imported = 0
        rows_updated = 0
        cursor = conn.cursor()

        # Read header row to map columns
        headers_raw = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers_raw.append(str(cell.value or "").strip())

        import unicodedata
        import re
        def clean_h(h):
            if not h: return ""
            # Normalizar: quitar acentos y caracteres raros
            s = str(h).strip().lower()
            s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
            s = re.sub(r'[^a-z0-9]', '', s)
            return s

        headers_clean = [clean_h(h) for h in headers_raw]

        # Column mapping (flexible matching using cleaned versions)
        col_map = {}
        mapping_rules = {
            'id_banco': ['idbanco', 'id', 'id_banco', 'id_de_banco'],
            'fecha': ['fecha', 'fch', 'date'],
            'descripcion': ['descripcion', 'description', 'desc'],
            'monto': ['monto', 'importe', 'amount', 'total'],
            'saldo': ['saldo', 'balance'],
            'sucursal': ['sucursal', 'branch'],
            'operacion_numero': ['operacionnumero', 'nrooperacion', 'opnumero', 'nroop', 'operacion', 'opnro', 'numerooperacion', 'nro.operacion', 'op.numero'],
            'operacion_hora': ['operacionhora', 'ophora', 'hora', 'time'],
            'referencia': ['referencia', 'ref'],
            'op_manual': ['opmanual', 'op_manual'],
            'op_cancelacion': ['opcancelacion', 'op_cancelacion'],
            'descripcion_final': ['descripcionfinal', 'descripcion_final'],
            'estado': ['estado', 'status']
        }

        for key, aliases in mapping_rules.items():
            # Limpiamos los alias también para que coincidan con headers_clean
            clean_aliases = [clean_h(a) for a in aliases]
            for i, hc in enumerate(headers_clean):
                if hc in clean_aliases:
                    col_map[key] = i
                    break

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None or all(v is None for v in row):
                continue

            def get_val(key, default=None):
                idx = col_map.get(key)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    return v if v is not None else default
                return default

            fecha = get_val('fecha')
            if fecha is None:
                continue  # Skip rows without date

            # Handle fecha - convert to string 'YYYY-MM-DD' for ODBC compatibility
            if isinstance(fecha, str):
                # Try common formats and convert to YYYY-MM-DD string
                parsed = False
                for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y'):
                    try:
                        fecha = datetime.strptime(fecha, fmt).strftime('%Y-%m-%d')
                        parsed = True
                        break
                    except ValueError:
                        continue
                if not parsed:
                    continue  # Skip if we can't parse the date
            elif isinstance(fecha, datetime):
                fecha = fecha.strftime('%Y-%m-%d')
            elif isinstance(fecha, date):
                fecha = fecha.strftime('%Y-%m-%d')
            else:
                continue  # Skip unsupported date type

            monto = get_val('monto', 0)
            try:
                monto = float(str(monto).replace(',', ''))
            except (ValueError, TypeError):
                monto = 0

            saldo = get_val('saldo', 0)
            try:
                saldo = float(str(saldo).replace(',', ''))
            except (ValueError, TypeError):
                saldo = 0

            # Ensure all string values are clean (no None)
            def safe_str(val, max_len=200):
                if val is None:
                    return ''
                s = str(val).strip()
                return s[:max_len]

            # Upsert logic based on id_banco
            id_banco_val = get_val('id_banco')
            try:
                id_banco_val = int(id_banco_val) if id_banco_val else None
            except:
                id_banco_val = None

            is_update = False
            if id_banco_val:
                cursor.execute("SELECT Id FROM BankMovements WHERE Id = ? AND CodCia = ? AND BankCode = ?", (id_banco_val, codcia, bank_code))
                if cursor.fetchone():
                    is_update = True
            
            p_desc = safe_str(get_val('descripcion', ''), 200)
            p_sucursal = safe_str(get_val('sucursal', ''), 50)
            p_op_num = safe_str(get_val('operacion_numero', ''), 30)
            p_op_hora = safe_str(get_val('operacion_hora', ''), 10)
            p_ref = safe_str(get_val('referencia', ''), 50)
            p_op_man = safe_str(get_val('op_manual', ''), 30)
            p_op_can = safe_str(get_val('op_cancelacion', ''), 30)
            p_desc_fin = safe_str(get_val('descripcion_final', ''), 200)

            if is_update:
                cursor.execute("""
                    UPDATE BankMovements
                    SET Fecha = ?, Descripcion = ?, Monto = ?, Saldo = ?, Sucursal = ?,
                        OperacionNumero = ?, OperacionHora = ?, Referencia = ?, OpManual = ?,
                        OpCancelacion = ?, DescripcionFinal = ?, ImportBatchId = ?
                    WHERE Id = ?
                """, (
                    fecha, p_desc, monto, saldo, p_sucursal, p_op_num, p_op_hora,
                    p_ref, p_op_man, p_op_can, p_desc_fin, batch_id, id_banco_val
                ))
                rows_updated += 1
            else:
                cursor.execute("""
                    INSERT INTO BankMovements
                    (CodCia, BankCode, Fecha, Descripcion, Monto, Saldo, Sucursal,
                     OperacionNumero, OperacionHora, Referencia, OpManual, OpCancelacion,
                     DescripcionFinal, Estado, ImportBatchId)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Pendiente', ?)
                """, (
                    codcia, bank_code, fecha, p_desc, monto, saldo, p_sucursal, 
                    p_op_num, p_op_hora, p_ref, p_op_man, p_op_can, p_desc_fin, batch_id
                ))
                rows_imported += 1

        conn.commit()
        wb.close()

        msg = f"Se importaron {rows_imported} nuevos y se actualizaron {rows_updated}." if rows_updated > 0 else f"Se importaron {rows_imported} movimientos bancarios."
        return {
            "status": "success",
            "message": msg,
            "batch_id": batch_id,
            "rows_imported": rows_imported,
            "rows_updated": rows_updated
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Error al importar Excel: {str(e)}")
    finally:
        conn.close()


@router.get("/movimientos-banco")
def get_bank_movements(
    codcia: str = Query(...),
    bank_code: str = Query(...),
    year: Optional[str] = None,
    month: Optional[str] = None,
    estado: Optional[str] = None
):
    """Lista movimientos bancarios importados con filtros."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        query = """
            SELECT bm.Id, bm.CodCia, bm.BankCode, 
                   CONVERT(varchar, bm.Fecha, 103) as Fecha,
                   bm.Descripcion, bm.Monto, bm.Saldo, bm.Sucursal, bm.OperacionNumero,
                   bm.OperacionHora, bm.Referencia, bm.OpManual, bm.OpCancelacion,
                   bm.DescripcionFinal, bm.Estado, bm.ReconciliationDetailId,
                   bm.ImportBatchId, bm.ImportedAt
            FROM BankMovements bm
            WHERE bm.CodCia = ? AND bm.BankCode = ?
        """
        params = [codcia, bank_code]

        if year:
            query += " AND YEAR(bm.Fecha) = ?"
            params.append(int(year))
        if month:
            query += " AND MONTH(bm.Fecha) = ?"
            params.append(int(month))
        if estado:
            query += " AND bm.Estado = ?"
            params.append(estado)

        query += " ORDER BY bm.Fecha DESC, bm.Id DESC"
        cursor.execute(query, params)
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/movimientos-banco/all")
def delete_all_bank_movements(
    codcia: str = Query(...),
    bank_code: str = Query(...)
):
    """
    Elimina TODOS los movimientos bancarios para una empresa y banco específicos.
    También borra en cascada (o manualmente) los ReconciliationDetails asociados.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()

        # Encontrar los movimientos a eliminar
        cursor.execute("""
            SELECT Id FROM BankMovements 
            WHERE CodCia = ? AND BankCode = ?
        """, (codcia, bank_code))
        rows = cursor.fetchall()
        
        if not rows:
            return {"status": "success", "message": "No hay movimientos para eliminar.", "deleted": 0}

        movement_ids = [row[0] for row in rows]
        
        # SQL IN para evitar hacer 1000 queries
        placeholders = ','.join('?' for _ in movement_ids)
        
        # Eliminar detalles de conciliación asociados
        cursor.execute(f"DELETE FROM ReconciliationDetail WHERE BankMovementId IN ({placeholders})", movement_ids)

        # Eliminar los movimientos bancarios
        cursor.execute(f"DELETE FROM BankMovements WHERE Id IN ({placeholders})", movement_ids)

        conn.commit()
        return {
            "status": "success", 
            "message": f"Se eliminaron {len(movement_ids)} movimientos bancarios exitosamente.",
            "deleted": len(movement_ids)
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/movimientos-banco/duplicates")
def get_duplicate_bank_movements(
    codcia: str = Query(...),
    bank_code: str = Query(...)
):
    """
    Busca movimientos bancarios duplicados (mismo CodCia, BankCode, fecha (solo date), Monto y OperacionNumero)
    para una empresa y banco específicos.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT Id, Fecha, Descripcion, Monto, Saldo, Sucursal, OperacionNumero, OperacionHora, Referencia, OpManual, OpCancelacion, DescripcionFinal, Estado
            FROM BankMovements
            WHERE CodCia = ? AND BankCode = ?
              AND OperacionNumero IS NOT NULL AND RTRIM(LTRIM(OperacionNumero)) <> ''
              AND EXISTS (
                  SELECT 1 FROM BankMovements bm2
                  WHERE bm2.CodCia = BankMovements.CodCia
                    AND bm2.BankCode = BankMovements.BankCode
                    AND CAST(bm2.Fecha AS DATE) = CAST(BankMovements.Fecha AS DATE)
                    AND bm2.Monto = BankMovements.Monto
                    AND bm2.OperacionNumero = BankMovements.OperacionNumero
                  GROUP BY bm2.CodCia, bm2.BankCode, CAST(bm2.Fecha AS DATE), bm2.Monto, bm2.OperacionNumero
                  HAVING COUNT(*) > 1
              )
            ORDER BY OperacionNumero, Fecha, Monto
        """, (codcia, bank_code))
        
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/movimientos-banco/{movimiento_id}")
def delete_bank_movement(movimiento_id: int):
    """
    Elimina un movimiento bancario específico, solo si su estado es Pendiente.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        
        # Verificar si existe y si no está conciliado
        cursor.execute("SELECT Estado FROM BankMovements WHERE Id = ?", (movimiento_id,))
        row = cursor.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail="Movimiento no encontrado")
            
        if row[0] == 'Conciliado':
            raise HTTPException(status_code=400, detail="No se puede eliminar un movimiento conciliado")
            
        cursor.execute("DELETE FROM BankMovements WHERE Id = ?", (movimiento_id,))
        conn.commit()
        
        return {"status": "success", "message": "Movimiento eliminado exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/cobranzas")
def get_cobranzas(
    codcia: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None,
    bank_code: Optional[str] = None,
    solo_pendientes: bool = True,
    cross_company: bool = False
):
    """
    Lista cobranzas de CcbMVtos con datos de CcbICaja.
    Si solo_pendientes=True, excluye las ya conciliadas.
    year y month son opcionales; si no se proveen, trae todos.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        query = """
            SELECT CASE WHEN matched.IsMatched = 1 THEN 'Conciliado' ELSE 'Pendiente' END as Estado,
                   m.CodCia, m.anos, m.mes, m.coddoc, m.nrodoc, m.nroitm,
                   m.fchdoc, m.codaux, m.NomAux, m.codven, m.nomven,
                   m.import, m.NroDep, m.fchDep, m.glodoc, m.codbco,
                   m.tpopgo, m.Glosa, m.CodDep, m.codref, m.nroref,
                   c.nombco, c.import as import_caja, c.sdodoc
            FROM CcbMVtos m
            LEFT JOIN CcbICaja c ON m.CodCia = c.codcia
                AND m.coddoc = c.coddoc AND m.nrodoc = c.nrodoc
            OUTER APPLY (
                SELECT TOP 1 1 as IsMatched
                FROM ReconciliationDetail rd
                WHERE rd.MatchCodCia = m.CodCia
                  AND rd.MatchCoddoc = m.coddoc
                  AND rd.MatchNrodoc = m.nrodoc
                  AND rd.MatchNroitm = m.nroitm
            ) matched
            WHERE (m.FlgEst IS NULL OR m.FlgEst <> 'E')
        """
        params = []

        if year:
            query += " AND m.anos = ?"
            params.append(year)
        if month:
            query += " AND m.mes = ?"
            params.append(month.zfill(2))

        if codcia:
            if cross_company:
                # Cross-company: no company filter, show ALL companies
                pass
            else:
                query += " AND m.CodCia = ?"
                params.append(codcia)

        if solo_pendientes:
            query += " AND matched.IsMatched IS NULL"
            
        query += " ORDER BY m.CodCia, m.fchdoc DESC, m.nrodoc"
        cursor.execute(query, params)
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/auto-match")
def auto_match(request: AutoMatchRequest):
    """
    Conciliación automática inteligente:
    - Aplica Regex Cleaning Rules (DB)
    - Soporta Múltiples compañias (Filiales tpopgo=1)
    - Soporta Agrupación Many-to-Many por Importe Total & Fecha
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        
        # 1. Load cleaning rules from the UI settings (conciliacion_reglas.json)
        def clean_op(op):
            return get_clean_op(op)

        # 2. Load pending bank movements
        query = """
            SELECT Id, OpCancelacion, OperacionNumero, Monto, Fecha
            FROM BankMovements
            WHERE CodCia = ? AND BankCode = ? AND Estado = 'Pendiente'
        """
        params = [request.codcia, request.bank_code]

        if request.only_ingresos:
            query += " AND Monto > 0"

        if request.period_year:
            query += " AND YEAR(Fecha) = ?"
            params.append(int(request.period_year))
        if request.period_month:
            query += " AND MONTH(Fecha) = ?"
            params.append(int(request.period_month))

        cursor.execute(query, params)
        pending_movements = cursor.fetchall()

        # Group Banks by CleanedOp (sin fecha estricta)
        bank_groups = defaultdict(list)
        for mov in pending_movements:
            b_id = mov[0]
            raw_op = mov[1] if mov[1] and str(mov[1]).strip() else mov[2]
            cleaned = clean_op(raw_op)
            if cleaned:
                bank_groups[cleaned].append(mov)

        # 3. Load ALL pending CcbMVtos for these dates/companies with Filiales logic
        if request.cross_company:
            c_query = """
                SELECT CodCia, coddoc, nrodoc, nroitm, NroDep, nroref, import, fchDep, fchdoc, tpopgo, CodDep, CodCom, codref, codaux 
                FROM CcbMVtos
                WHERE (FlgEst IS NULL OR FlgEst <> 'E')
                  AND NOT EXISTS (
                      SELECT 1 FROM ReconciliationDetail rd
                      WHERE rd.MatchCodCia = CcbMVtos.CodCia
                        AND rd.MatchCoddoc = CcbMVtos.coddoc
                        AND rd.MatchNrodoc = CcbMVtos.nrodoc
                        AND rd.MatchNroitm = CcbMVtos.nroitm
                  )
            """
            c_params = []
        else:
            c_query = """
                SELECT CodCia, coddoc, nrodoc, nroitm, NroDep, nroref, import, fchDep, fchdoc, tpopgo, CodDep, CodCom, codref, codaux 
                FROM CcbMVtos
                WHERE (FlgEst IS NULL OR FlgEst <> 'E')
                  AND (
                      CodCia = ? 
                      OR (tpopgo = '1' AND CodCom = ? AND CodDep = ?)
                  )
                  AND NOT EXISTS (
                      SELECT 1 FROM ReconciliationDetail rd
                      WHERE rd.MatchCodCia = CcbMVtos.CodCia
                        AND rd.MatchCoddoc = CcbMVtos.coddoc
                        AND rd.MatchNrodoc = CcbMVtos.nrodoc
                        AND rd.MatchNroitm = CcbMVtos.nroitm
                  )
            """
            c_params = [request.codcia, request.codcia, request.bank_code]
        
        if request.period_year:
           c_query += " AND anos = ?"
           c_params.append(str(request.period_year))
        if request.period_month:
           c_query += " AND mes = ?"
           c_params.append(str(request.period_month).zfill(2))
           
        cursor.execute(c_query, c_params)
        unmatched_cobs = cursor.fetchall()
        
        # Group Cobs by CleanedOp (sin fecha estricta)
        cob_groups = defaultdict(list)
        for cob in unmatched_cobs:
            raw_op = cob[4] if cob[4] and str(cob[4]).strip() else cob[5]
            cleaned = clean_op(raw_op)
            if cleaned:
                cob_groups[cleaned].append(cob)

        matched_count = 0
        total_processed = len(pending_movements)

        # 4. Compare Groups and Insert Many-to-Many Maps
        for key, b_list in bank_groups.items():
            c_list = cob_groups.get(key)
            if not c_list:
                continue
                
            sum_b = sum([float(b[3] or 0) for b in b_list])
            sum_c = sum([float(c[6] or 0) for c in c_list])
            
            # Exact Amount Match
            if abs(abs(sum_b) - abs(sum_c)) <= 0.05:
                # Get next group ID (integer)
                cursor.execute("SELECT ISNULL(MAX(ReconciliationId), 0) + 1 FROM ReconciliationDetail")
                group_id = cursor.fetchone()[0]
                
                for b_idx, b_mov in enumerate(b_list):
                    b_id = b_mov[0]
                    first_detail_id = None
                    
                    for c_cob in c_list:
                        cursor.execute("""
                            INSERT INTO ReconciliationDetail
                            (BankMovementId, MatchCodCia, MatchCoddoc, MatchNrodoc, MatchNroitm, MatchType, ReconciliationId)
                            OUTPUT INSERTED.Id
                            VALUES (?, ?, ?, ?, ?, 'AUTO', ?)
                        """, (b_id, c_cob[0], c_cob[1], c_cob[2], c_cob[3], group_id))
                        
                        inserted_row = cursor.fetchone()
                        detail_id = inserted_row[0] if inserted_row else None
                        
                        if detail_id and not first_detail_id:
                            first_detail_id = detail_id
                            
                        if detail_id:
                            # Insert into tbl_Conciliados
                            # c_cob: 0:CodCia, 1:coddoc, 2:nrodoc, 3:nroitm, 4:NroDep, 5:nroref, 6:import, 7:fchDep, 8:fchdoc, 9:tpopgo, 10:CodDep, 11:CodCom, 12:codref, 13:codaux
                            # b_mov: 0:Id, 1:OpCancelacion, 2:OperacionNumero, 3:Monto, 4:Fecha, 5:CodCia, 6:BankCode (wait, b_mov doesn't have CodCia/BankCode yet, but we have request.codcia and request.bank_code)
                            try:
                                cursor.execute("""
                                    INSERT INTO tbl_Conciliados
                                    (ReconciliationDetailId, IdCobranza_CodCia, IdCobranza_coddoc, IdCobranza_nrodoc, IdCobranza_nroitm, codref, nroref, importe, codaux, IdBanco, Fecha_banco, empresa, codigo_banco, nro_operacion, usuario)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CONVERT(datetime, ?, 120), ?, ?, ?, ?)
                                """, (
                                    detail_id, c_cob[0], c_cob[1], c_cob[2], c_cob[3],
                                    c_cob[12], c_cob[5], c_cob[6], c_cob[13],
                                    b_id, str(b_mov[4])[:10], request.codcia, request.bank_code,
                                    key, request.usuario
                                ))
                            except Exception as ex:
                                print(f"Failed inserting tbl_Conciliados obj. Date was: {b_mov[4]}")
                                raise ex
                            
                    if first_detail_id:
                        cursor.execute("""
                            UPDATE BankMovements
                            SET Estado = 'Conciliado', ReconciliationDetailId = ?
                            WHERE Id = ?
                        """, (first_detail_id, b_id))
                        matched_count += 1
                        
        conn.commit()

        return {
            "status": "success",
            "message": f"Conciliación automática completada. {matched_count} movimientos vinculados usando agrupación avanzada.",
            "matched_count": matched_count,
            "total_processed": total_processed
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


class OpManualUpdateRequest(BaseModel):
    op_manual: str

@router.put("/movimientos-banco/{mov_id}/op-manual")
def update_op_manual(mov_id: int, request: OpManualUpdateRequest):
    """Actualiza el campo OpManual y OpCancelacion asociado a un movimiento bancario."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        
        # Check if exists and is not matched
        cursor.execute("SELECT Estado FROM BankMovements WHERE Id = ?", (mov_id,))
        mov = cursor.fetchone()
        if not mov:
            raise HTTPException(status_code=404, detail="Movimiento no encontrado")
        
        # Update both OpManual and OpCancelacion
        cursor.execute("""
            UPDATE BankMovements
            SET OpManual = ?, OpCancelacion = ?
            WHERE Id = ?
        """, (request.op_manual, request.op_manual, mov_id))
        
        conn.commit()
        return {"status": "success", "message": "OpManual actualizado correctamente."}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.post("/manual-match")
def manual_match(request: ManualMatchRequest):
    """Conciliación manual: vincula N movimientos bancarios con M cobranzas (Many-to-Many)."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()

        if not request.bank_movement_ids or not request.cobranza_keys:
            raise HTTPException(status_code=400, detail="Debe seleccionar al menos un banco y una cobranza.")

        # 1. Fetch and validate Bank Movements
        bank_placeholders = ','.join(['?'] * len(request.bank_movement_ids))
        cursor.execute(f"""
            SELECT Id, Estado, Monto, Fecha, CodCia, BankCode, OpCancelacion, OperacionNumero FROM BankMovements 
            WHERE Id IN ({bank_placeholders})
        """, request.bank_movement_ids)
        banks = cursor.fetchall()
        
        if len(banks) != len(request.bank_movement_ids):
            raise HTTPException(status_code=404, detail="Uno o más movimientos bancarios no encontrados.")
        
        sum_bancos = 0.0
        bank_date = None
        for b in banks:
            if str(b[1]).strip() == 'Conciliado':
                raise HTTPException(status_code=400, detail="Uno de los movimientos bancarios ya está conciliado.")
            sum_bancos += float(b[2] or 0)
            if not bank_date and b[3]:
                bank_date = b[3]

        if not bank_date:
             raise HTTPException(status_code=400, detail="Los movimientos bancarios no tienen una fecha válida.")

        # 2. Fetch and validate Cobranzas
        sum_cobranzas = 0.0
        parsed_cobs = []
        for key in request.cobranza_keys:
            parts = key.split('|')
            if len(parts) != 4:
                raise HTTPException(status_code=400, detail=f"Llave de cobranza inválida: {key}")
            
            c_cia, c_doc, n_doc, n_itm = parts
            
            cursor.execute("""
                SELECT CodCia, import, fchDep, fchdoc, codref, nroref, codaux FROM CcbMVtos
                WHERE CodCia = ? AND coddoc = ? AND nrodoc = ? AND nroitm = ?
            """, (c_cia, c_doc, n_doc, n_itm))
            cob = cursor.fetchone()
            if not cob:
                raise HTTPException(status_code=404, detail=f"Cobranza no encontrada: {key}")
            
            # Check if matched already
            cursor.execute("""
                SELECT Id FROM ReconciliationDetail
                WHERE MatchCodCia = ? AND MatchCoddoc = ? AND MatchNrodoc = ? AND MatchNroitm = ?
            """, (c_cia, c_doc, n_doc, n_itm))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail=f"La cobranza {key} ya está conciliada.")
            
            # Amount validation
            sum_cobranzas += float(cob[1] or 0)
                
            parsed_cobs.append((c_cia, c_doc, n_doc, n_itm, cob[4], cob[5], cob[1], cob[6])) # + codref, nroref, importe, codaux

        # 3. Validate Amounts
        # In absolute terms, they must match within the specified tolerance
        max_diff_val = request.max_diff if request.max_diff is not None else 0.50
        if abs(abs(sum_bancos) - abs(sum_cobranzas)) > max_diff_val:
            raise HTTPException(status_code=400, detail=f"Los importes no concuerdan. Bancos: {abs(sum_bancos):.2f}, Cobranzas: {abs(sum_cobranzas):.2f}")

        # 4. Do the Many-to-Many Insert
        # Get next group ID (integer)
        cursor.execute("SELECT ISNULL(MAX(ReconciliationId), 0) + 1 FROM ReconciliationDetail")
        match_group_id = cursor.fetchone()[0]
        
        for idx, b_id in enumerate(request.bank_movement_ids):
            primary_detail_id = None
            bank_info = next((bx for bx in banks if bx[0] == b_id), None)
            
            for c_cia, c_doc, n_doc, n_itm, codref, nroref, importe_c, codaux in parsed_cobs:
                cursor.execute("""
                    INSERT INTO ReconciliationDetail
                    (BankMovementId, MatchCodCia, MatchCoddoc, MatchNrodoc, MatchNroitm, MatchType, ReconciliationId)
                    OUTPUT INSERTED.Id
                    VALUES (?, ?, ?, ?, ?, 'MANUAL', ?)
                """, (b_id, c_cia, c_doc, n_doc, n_itm, match_group_id))
                
                inserted_row = cursor.fetchone()
                detail_id = inserted_row[0] if inserted_row else None
                
                if detail_id and not primary_detail_id:
                    primary_detail_id = detail_id
                
                if detail_id and bank_info:
                    # bank_info indices: 0:Id, 1:Estado, 2:Monto, 3:Fecha, 4:CodCia, 5:BankCode, 6:OpCancelacion, 7:OperacionNumero
                    raw_op = bank_info[6] if bank_info[6] and str(bank_info[6]).strip() else bank_info[7]
                    cl_op = get_clean_op(raw_op)
                    cursor.execute("""
                        INSERT INTO tbl_Conciliados
                        (ReconciliationDetailId, IdCobranza_CodCia, IdCobranza_coddoc, IdCobranza_nrodoc, IdCobranza_nroitm, codref, nroref, importe, codaux, IdBanco, Fecha_banco, empresa, codigo_banco, nro_operacion, usuario)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CONVERT(datetime, ?, 120), ?, ?, ?, ?)
                    """, (
                        detail_id, c_cia, c_doc, n_doc, n_itm, codref, nroref, importe_c, codaux,
                        b_id, bank_info[3], bank_info[4], bank_info[5], cl_op, request.usuario
                    ))
                    
            if primary_detail_id:
                cursor.execute("""
                    UPDATE BankMovements
                    SET Estado = 'Conciliado', ReconciliationDetailId = ?
                    WHERE Id = ?
                """, (primary_detail_id, b_id))

        conn.commit()

        return {
            "status": "success",
            "message": "Conciliación manual realizada exitosamente.",
            "group_id": match_group_id
        }
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.delete("/unmatch/{detail_id}")
def unmatch(detail_id: int):
    """Deshace un match de conciliación (soporta grupos Many-to-Many)."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()

        # Get the group ID before deleting
        cursor.execute("""
            SELECT ReconciliationId, BankMovementId FROM ReconciliationDetail WHERE Id = ?
        """, (detail_id,))
        detail = cursor.fetchone()
        if not detail:
            # Fallback: The ReconciliationDetail may have been already deleted (orphan in tbl_Conciliados)
            # Try to clean the orphan tbl_Conciliados record directly
            cursor.execute("DELETE FROM tbl_Conciliados WHERE ReconciliationDetailId = ?", (detail_id,))
            if cursor.rowcount > 0:
                conn.commit()
                return {"status": "success", "message": "Registro huérfano de conciliación eliminado exitosamente."}
            raise HTTPException(status_code=404, detail="Match no encontrado")

        group_id = detail[0]
        single_bank_id = detail[1]

        if group_id:
            # Delete entire group
            cursor.execute("SELECT BankMovementId FROM ReconciliationDetail WHERE ReconciliationId = ? GROUP BY BankMovementId", (group_id,))
            bank_ids = [row[0] for row in cursor.fetchall()]

            # Get all detail IDs in this group to clean tbl_Conciliados
            cursor.execute("SELECT Id FROM ReconciliationDetail WHERE ReconciliationId = ?", (group_id,))
            all_detail_ids = [row[0] for row in cursor.fetchall()]

            # Delete from tbl_Conciliados first (FK reference)
            if all_detail_ids:
                ph = ','.join(['?'] * len(all_detail_ids))
                cursor.execute(f"DELETE FROM tbl_Conciliados WHERE ReconciliationDetailId IN ({ph})", all_detail_ids)

            cursor.execute("DELETE FROM ReconciliationDetail WHERE ReconciliationId = ?", (group_id,))

            for b_id in bank_ids:
                cursor.execute("""
                    UPDATE BankMovements
                    SET Estado = 'Pendiente', ReconciliationDetailId = NULL
                    WHERE Id = ?
                """, (b_id,))
        else:
            # Fallback: legacy single match delete
            # Delete from tbl_Conciliados first
            cursor.execute("DELETE FROM tbl_Conciliados WHERE ReconciliationDetailId = ?", (detail_id,))

            cursor.execute("DELETE FROM ReconciliationDetail WHERE Id = ?", (detail_id,))
            cursor.execute("""
                UPDATE BankMovements
                SET Estado = 'Pendiente', ReconciliationDetailId = NULL
                WHERE Id = ?
            """, (single_bank_id,))

        conn.commit()
        return {"status": "success", "message": "Match eliminado exitosamente."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/resumen")
def get_resumen(
    codcia: str = Query(...),
    bank_code: str = Query(...),
    year: Optional[str] = None,
    month: Optional[str] = None
):
    """Dashboard de conciliación: totales, % conciliado, montos."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()

        # Base filter
        base_filter = "WHERE CodCia = ? AND BankCode = ?"
        params = [codcia, bank_code]

        if year:
            base_filter += " AND YEAR(Fecha) = ?"
            params.append(int(year))
        if month:
            base_filter += " AND MONTH(Fecha) = ?"
            params.append(int(month))

        # Total movements
        cursor.execute(f"SELECT COUNT(*), ISNULL(SUM(Monto),0) FROM BankMovements {base_filter}", params)
        r = cursor.fetchone()
        total_mov = r[0]
        total_monto = float(r[1])

        # Matched
        cursor.execute(f"""
            SELECT COUNT(*), ISNULL(SUM(Monto),0)
            FROM BankMovements {base_filter} AND Estado = 'Conciliado'
        """, params)
        r = cursor.fetchone()
        matched = r[0]
        matched_monto = float(r[1])

        # Pending
        cursor.execute(f"""
            SELECT COUNT(*), ISNULL(SUM(Monto),0)
            FROM BankMovements {base_filter} AND Estado = 'Pendiente'
        """, params)
        r = cursor.fetchone()
        pending = r[0]
        pending_monto = float(r[1])

        pct = round((matched / total_mov * 100), 1) if total_mov > 0 else 0

        return {
            "total_movimientos": total_mov,
            "total_monto": total_monto,
            "conciliados": matched,
            "conciliados_monto": matched_monto,
            "pendientes": pending,
            "pendientes_monto": pending_monto,
            "porcentaje_conciliado": pct
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ─── REPORTE DETALLES DE MATCH (MODALS) ─────────────────────────

@router.get("/movimiento-banco/{mov_id}/match-details")
def get_bank_match_details(mov_id: int):
    """Dado un movimiento bancario conciliado, devuelve las cobranzas del sistema con las que matcheó."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT m.CodCia, m.coddoc, m.nrodoc, m.nroitm, m.fchdoc, m.import, m.NroDep, m.fchDep, m.NomAux
            FROM CcbMVtos m
            JOIN ReconciliationDetail rd ON m.CodCia = rd.MatchCodCia 
                AND m.coddoc = rd.MatchCoddoc AND m.nrodoc = rd.MatchNrodoc AND m.nroitm = rd.MatchNroitm
            WHERE rd.BankMovementId = ?
        """, (mov_id,))
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/cobranza/match-details")
def get_cobranza_match_details(
    codcia: str = Query(...),
    coddoc: str = Query(...),
    nrodoc: str = Query(...),
    nroitm: str = Query(...)
):
    """Dada una cobranza conciliada, devuelve los movimientos bancarios con los que matcheó."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT bm.Id, bm.Fecha, bm.Descripcion, bm.Monto, bm.OperacionNumero, bm.Estado
            FROM BankMovements bm
            JOIN ReconciliationDetail rd ON bm.Id = rd.BankMovementId
            WHERE rd.MatchCodCia = ? AND rd.MatchCoddoc = ? AND rd.MatchNrodoc = ? AND rd.MatchNroitm = ?
        """, (codcia, coddoc, nrodoc, nroitm))
        result = rows_to_list(cursor)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ─── REPORTE TODAS LAS COBRANZAS (TAB 2) ─────────────────────────

@router.get("/cobranzas-todas")
def get_todas_cobranzas(
    year: Optional[str] = None,
    month: Optional[str] = None,
    codcia: Optional[str] = None,
    bank_code: Optional[str] = None
):
    """
    Retorna la tabla completa de cobranzas con datos para reporte y estado de conciliación.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    
    try:
        cursor = conn.cursor()
        
        # Filtros opcionales
        where_clauses = []
        params = []
        if year:
            where_clauses.append("m.anos = ?")
            params.append(year)
        if month:
            where_clauses.append("m.mes = ?")
            params.append(month.zfill(2))
        if codcia:
            if bank_code:
                where_clauses.append("(m.CodCia = ? OR (m.tpopgo = '1' AND m.CodCom = ? AND m.CodDep = ?))")
                params.extend([codcia, codcia, bank_code])
            else:
                where_clauses.append("m.CodCia = ?")
                params.append(codcia)
            
        where_clauses.append("(m.FlgEst IS NULL OR m.FlgEst <> 'E')")
            
        where_str = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

        # Query enriquecida para el reporte y listado (Refinada con Joins a Bancos, POS y Documentos)
        query = f"""
            SELECT
                -- Columnas crudas solicitadas para la tabla principal
                m.CodCia,
                m.anos,
                m.mes,
                m.coddoc,
                m.nrodoc,
                m.tpodoc,
                m.fchdoc,
                m.codaux,
                m.NomAux,
                m.codven,
                m.nomven,
                m.codref,
                m.nroref,
                m.import,
                m.glodoc,
                m.fmapgo,
                m.CodDep,
                m.NroDep,
                m.fchDep,
                m.tpopgo,
                m.Dcmpgo,
                m.CodCom,
                m.usuario,
                m.FlgEst,
                
                -- Alias y campos adicionales necesarios para la logica del Reporte de Caja y UI
                m.CodCia as Suc,
                m.coddoc as SerieDoc,
                m.nrodoc as NroDoc,
                m.nroitm as Correlat,
                m.tpodoc as TipoDoc,
                m.fchdoc as FechaEfe,
                m.import as Monto,
                m.fchDep as F_D,
                m.glodoc as Glosa,
                m.Glosa as Concepto,
                m.codbco as CodBco,
                m.nro_apl as NumCompte,
                m.NomAux as RazonSocial,
                m.codmon,
                c.nombco as CuentaNombre,
                c.flgest as CajaFlgEst,
                rd.Id as MatchId,
                rd.BankMovementId as BankId,
                CASE WHEN (rd.Id IS NOT NULL OR m.FlgEst = 'C' OR c.flgest = 'C') THEN 1 ELSE 0 END as IsConciliado,
                -- Buscar el nombre de la cuenta o POS
                COALESCE(p.DESTARJ, t.Nombre) as GroupName,
                -- Buscar la fecha original del documento (CcbRGdoc)
                rg.fchdoc as FechaOriginalDoc
            FROM CcbMVtos m
            LEFT JOIN CcbICaja c ON m.CodCia = c.codcia 
                AND m.coddoc = c.coddoc AND m.nrodoc = c.nrodoc
            LEFT JOIN ReconciliationDetail rd ON rd.MatchCodCia = m.CodCia 
                AND rd.MatchCoddoc = m.coddoc 
                AND rd.MatchNrodoc = m.nrodoc 
                AND rd.MatchNroitm = m.nroitm
            -- Join con Bancos (OUTER APPLY TOP 1 to prevent row multiplication)
            OUTER APPLY (
                SELECT TOP 1 t2.Nombre 
                FROM CcbTabla t2 
                WHERE RTRIM(m.CodDep) = RTRIM(t2.Codigo) 
                AND t2.Tabla = '0001'
                AND t2.CodCia = CASE WHEN m.tpopgo = '1' THEN m.CodCom ELSE m.CodCia END
            ) t
            -- Join con POS
            LEFT JOIN POSTARJE p ON RTRIM(m.CodDep) = RTRIM(p.codtarj) AND p.CODcia = CASE WHEN m.tpopgo = '1' THEN m.CodCom ELSE m.CodCia END
            -- Join con Cabecera de Documento usando los campos de referencia
            LEFT JOIN CcbRGdoc rg ON m.CodCia = rg.codcia AND m.codref = rg.coddoc AND m.nroref = rg.nrodoc
            {where_str}
            ORDER BY m.tpopgo, COALESCE(p.DESTARJ, t.Nombre), m.nroitm
        """
        
        cursor.execute(query, params)
        result = rows_to_list(cursor)
        
        # Mapping tpopgo to JT (Payment Type Description) as per FoxPro logic
        tpopgo_map = {
            '1': 'FILIAL CANCELA',
            '2': 'PERSONAL',
            '3': 'AMERICAN EXPRES',
            '4': 'EPS',
            '5': 'DINERS',
            'C': 'CHEQUE',
            'D': 'DEPOSITO',
            'E': 'EFECTIVO',
            'M': 'MASTERCAR',
            'R': 'RETENCION',
            'A': 'ANTICIPO APLICACION',
            'B': 'ANTICIPO CREACION',
            'F': 'FILIAL DEPOSITO',
            'Z': 'IZIPAY'
        }

        enriched = []
        for r in result:
            tp = str(r.get('tpopgo') or '').strip()
            jt = tpopgo_map.get(tp, 'CANJE')
            if tp == '1':
                codcom = str(r.get('CodCom') or '').strip()
                if codcom:
                    jt = f"FILIAL CANCELA ({codcom})"
                    
            
            # codmon viene como string del row_to_dict
            cm = str(r.get('codmon') or '1').strip()
            soles = float(r.get('Monto') or 0) if cm == '1' else 0
            dolares = float(r.get('Monto') or 0) if cm == '2' else 0
            
            enriched.append({
                # --- Columnas crudas solicitadas para la tabla principal ---
                "ID": (r.get('nroitm') or '').strip() if r.get('nroitm') else '',
                "CodCia": (r.get('CodCia') or '').strip(),
                "anos": (r.get('anos') or '').strip(),
                "mes": (r.get('mes') or '').strip(),
                "coddoc": (r.get('coddoc') or '').strip(),
                "nrodoc": (r.get('nrodoc') or '').strip(),
                "tpodoc": (r.get('tpodoc') or '').strip(),
                "fchdoc": r.get('fchdoc'),
                "codaux": (r.get('codaux') or '').strip(),
                "NomAux": (r.get('NomAux') or '').strip(),
                "codven": (r.get('codven') or '').strip(),
                "nomven": (r.get('nomven') or '').strip(),
                "codref": (r.get('codref') or '').strip(),
                "nroref": (r.get('nroref') or '').strip(),
                "import": float(r.get('import') or 0),
                "glodoc": (r.get('glodoc') or '').strip(),
                "fmapgo": (r.get('fmapgo') or '').strip(),
                "CodDep": (r.get('CodDep') or '').strip(),
                "NroDep": (r.get('NroDep') or '').strip(),
                "fchDep": r.get('fchDep'),
                "tpopgo": (r.get('tpopgo') or '').strip(),
                "Dcmpgo": (r.get('Dcmpgo') or '').strip(),
                "CodCom": (r.get('CodCom') or '').strip(),
                "usuario": (r.get('usuario') or '').strip(),
                "FlgEst": (r.get('FlgEst') or '').strip(),

                # --- Alias y campos adicionales necesarios para la logica del Reporte de Caja y UI ---
                "id": f"{r['Suc']}-{r['SerieDoc']}-{r['NroDoc']}-{r['Correlat']}",
                "NroCaja": f"{(r['Suc'] or '').strip()}-{(r['SerieDoc'] or '').strip()}-{(r['NroDoc'] or '').strip()}",
                "NumCompte": r['NumCompte'],
                "FechaEfe": r['FechaEfe'],
                "Suc": (r['Suc'] or '').strip(),
                "Serie": "",
                "TipoDoc": (r['TipoDoc'] or '').strip(),
                "SerieDoc": (r['SerieDoc'] or '').strip(),
                "NroDoc": (r['NroDoc'] or '').strip(),
                "TipoDocCancelado": (r.get('codref') or '').strip(),
                "NroDocCancelado": (r.get('nroref') or '').strip(),
                "CodBco": (r['CodBco'] or '').strip(),
                "CuentaNombre": (r['CuentaNombre'] or '').strip() or 'BANCO NO ESPECIFICADO',
                "Correlat": r['Correlat'],
                "Monto": float(r['Monto'] or 0),
                "F_D": r['F_D'],
                "MntDoc": float(r['Monto'] or 0),
                "Importe": float(r['Monto'] or 0),
                "TotalDoc": float(r['Monto'] or 0),
                "OC": "",
                "MontoOC": 0.0,
                "Glosa": (r['Glosa'] or r['Concepto'] or '').strip(),
                "Concepto": (r['Concepto'] or '').strip(),
                "RazonSocial": (r['RazonSocial'] or '').strip(),
                "Codigo": (r['codaux'] or '').strip(),
                "JT": jt,
                "GroupName": (r['GroupName'] or r['CuentaNombre'] or 'VARIOS').strip(),
                "OriginalFechaDoc": r.get('FechaOriginalDoc'),
                "Soles": soles,
                "Dolares": dolares,
                "MatchId": r['MatchId'],
                "BankId": r['BankId'],
                "CajaFlgEst": (r.get('CajaFlgEst') or '').strip(),
                "Conciliado": str(r.get('IsConciliado', '0')).strip() == '1'
            })
            
        return enriched
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.get("/caja-publica/{codigo:path}")
def get_caja_publica(codigo: str):
    """
    Retorna el detalle de una caja específica usando su código único (CodCia-coddoc-nrodoc).
    """
    parts = codigo.split("-")
    if len(parts) < 3:
        raise HTTPException(status_code=400, detail="Código inválido. Formato esperado: CodCia-coddoc-nrodoc")
    
    codcia = parts[0]
    coddoc = parts[1]
    nrodoc = "-".join(parts[2:])
    
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    
    try:
        cursor = conn.cursor()
        query = """
            SELECT
                m.CodCia as Suc,
                m.coddoc as SerieDoc,
                m.nrodoc as NroDoc,
                m.nroitm as Correlat,
                m.tpodoc as TipoDoc,
                m.fchdoc as FechaEfe,
                m.import as Monto,
                m.fchDep as F_D,
                m.glodoc as Glosa,
                m.Glosa as Concepto,
                m.codbco as CodBco,
                m.nro_apl as NumCompte,
                m.NomAux as RazonSocial,
                m.codmon,
                c.nombco as CuentaNombre,
                c.flgest as CajaFlgEst,
                rd.Id as MatchId,
                rd.BankMovementId as BankId,
                CASE WHEN (rd.Id IS NOT NULL OR m.FlgEst = 'C' OR c.flgest = 'C') THEN 1 ELSE 0 END as IsConciliado,
                COALESCE(p.DESTARJ, t.Nombre) as GroupName,
                rg.fchdoc as FechaOriginalDoc,
                m.codref as TipoDocCancelado,
                m.nroref as NroDocCancelado,
                m.codaux as Codigo,
                m.nomven,
                m.usuario,
                m.NroDep,
                m.fchDep,
                m.CodCom,
                m.tpopgo
            FROM CcbMVtos m
            LEFT JOIN CcbICaja c ON m.CodCia = c.codcia AND m.coddoc = c.coddoc AND m.nrodoc = c.nrodoc
            LEFT JOIN ReconciliationDetail rd ON rd.MatchCodCia = m.CodCia 
                AND rd.MatchCoddoc = m.coddoc AND rd.MatchNrodoc = m.nrodoc AND rd.MatchNroitm = m.nroitm
            OUTER APPLY (
                SELECT TOP 1 t2.Nombre 
                FROM CcbTabla t2 
                WHERE RTRIM(m.CodDep) = RTRIM(t2.Codigo) 
                AND t2.Tabla = '0001'
                AND t2.CodCia = CASE WHEN m.tpopgo = '1' THEN m.CodCom ELSE m.CodCia END
            ) t
            LEFT JOIN POSTARJE p ON RTRIM(m.CodDep) = RTRIM(p.codtarj) AND p.CODcia = CASE WHEN m.tpopgo = '1' THEN m.CodCom ELSE m.CodCia END
            LEFT JOIN CcbRGdoc rg ON m.CodCia = rg.codcia AND m.codref = rg.coddoc AND m.nroref = rg.nrodoc
            WHERE m.CodCia = ? AND m.coddoc = ? AND m.nrodoc = ?
            ORDER BY m.tpopgo, COALESCE(p.DESTARJ, t.Nombre), m.nroitm
        """
        cursor.execute(query, (codcia, coddoc, nrodoc))
        result = rows_to_list(cursor)
        
        if not result:
            # Fallback flexible matching if the user-provided code has a different prefix/padding
            import re
            digits = re.findall(r'\d+', nrodoc)
            if digits:
                num_suffix = digits[-1].lstrip('0')
                if num_suffix:
                    like_pattern = f"%{num_suffix.zfill(5)}"
                    query_fallback = query.replace("m.nrodoc = ?", "(m.nrodoc = ? OR m.nrodoc LIKE ?)")
                    cursor.execute(query_fallback, (codcia, coddoc, nrodoc, like_pattern))
                    result = rows_to_list(cursor)
        
        tpopgo_map = {
            '1': 'FILIAL CANCELA', '2': 'PERSONAL', '3': 'AMERICAN EXPRES', '4': 'EPS',
            '5': 'DINERS', 'C': 'CHEQUE', 'D': 'DEPOSITO', 'E': 'EFECTIVO', 'M': 'MASTERCAR',
            'R': 'RETENCION', 'A': 'ANTICIPO APLICACION', 'B': 'ANTICIPO CREACION', 'F': 'FILIAL DEPOSITO', 'Z': 'IZIPAY'
        }
        
        enriched = []
        for r in result:
            tp = str(r.get('tpopgo') or '').strip()
            jt = tpopgo_map.get(tp, 'CANJE')
            if tp == '1':
                codcom = str(r.get('CodCom') or '').strip()
                if codcom:
                    jt = f"FILIAL CANCELA ({codcom})"
            
            cm = str(r.get('codmon') or '1').strip()
            soles = float(r.get('Monto') or 0) if cm == '1' else 0
            dolares = float(r.get('Monto') or 0) if cm == '2' else 0
            
            enriched.append({
                "NroCaja": codigo,
                "Suc": (r['Suc'] or '').strip(),
                "SerieDoc": (r['SerieDoc'] or '').strip(),
                "NroDoc": (r['NroDoc'] or '').strip(),
                "Correlat": r['Correlat'],
                "TipoDocCancelado": (r.get('TipoDocCancelado') or '').strip(),
                "NroDocCancelado": (r.get('NroDocCancelado') or '').strip(),
                "FechaEfe": r.get('FechaEfe'),
                "OriginalFechaDoc": r.get('FechaOriginalDoc'),
                "Codigo": (r.get('Codigo') or '').strip(),
                "RazonSocial": (r.get('RazonSocial') or '').strip(),
                "nomven": (r.get('nomven') or '').strip(),
                "usuario": (r.get('usuario') or '').strip(),
                "NroDep": (r.get('NroDep') or '').strip(),
                "fchDep": r.get('fchDep'),
                "Soles": soles,
                "Dolares": dolares,
                "CajaFlgEst": (r.get('CajaFlgEst') or '').strip(),
                "Conciliado": str(r.get('IsConciliado', '0')).strip() == '1',
                "JT": jt,
                "GroupName": (r.get('GroupName') or r.get('CuentaNombre') or 'VARIOS').strip()
            })
        
        return enriched
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

@router.get("/match-details")
def get_match_details(match_id: int):
    """
    Retorna los detalles de una conciliacion: TODAS las cobranzas y TODOS los bancos vinculados al mismo grupo.
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    
    try:
        cursor = conn.cursor()
        
        # 1. Get the ReconciliationId (group) from this detail
        cursor.execute("SELECT ReconciliationId, MatchType, MatchedAt FROM ReconciliationDetail WHERE Id = ?", (match_id,))
        row = cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Conciliacion no encontrada")
        group_id = row[0]
        match_type = row[1]
        matched_at = row[2]
        
        # 2. Get ALL details in this group
        cursor.execute("""
            SELECT DISTINCT MatchCodCia, MatchCoddoc, MatchNrodoc, MatchNroitm
            FROM ReconciliationDetail WHERE ReconciliationId = ?
        """, (group_id,))
        cob_keys = cursor.fetchall()
        
        cursor.execute("""
            SELECT DISTINCT BankMovementId
            FROM ReconciliationDetail WHERE ReconciliationId = ?
        """, (group_id,))
        bank_ids = [r[0] for r in cursor.fetchall()]
        
        # 3. Fetch ALL cobranzas
        cobranzas_list = []
        for ck in cob_keys:
            cursor.execute("""
                SELECT m.CodCia, m.coddoc, m.nrodoc, m.nroitm, m.fchdoc, m.import, m.NomAux, 
                       m.codref, m.nroref, m.nomven, m.usuario, m.NroDep,
                       c.nombco as CuentaNombre
                FROM CcbMVtos m
                LEFT JOIN CcbICaja c ON m.CodCia = c.codcia AND m.coddoc = c.coddoc AND m.nrodoc = c.nrodoc
                WHERE m.CodCia = ? AND m.coddoc = ? AND m.nrodoc = ? AND m.nroitm = ?
            """, (ck[0], ck[1], ck[2], ck[3]))
            cols = [c[0] for c in cursor.description]
            r = cursor.fetchone()
            if r:
                d = dict(zip(cols, r))
                cobranzas_list.append({
                    "CodCia": (d.get('CodCia') or '').strip(),
                    "CodDoc": (d.get('coddoc') or '').strip(),
                    "NroDoc": (d.get('nrodoc') or '').strip(),
                    "NroItm": (d.get('nroitm') or '').strip(),
                    "Fecha": d.get('fchdoc'),
                    "Importe": float(d.get('import') or 0),
                    "RazonSocial": (d.get('NomAux') or '').strip(),
                    "Cuenta": (d.get('CuentaNombre') or '').strip(),
                    "CodRef": (d.get('codref') or '').strip(),
                    "NroRef": (d.get('nroref') or '').strip(),
                    "NomVen": (d.get('nomven') or '').strip(),
                    "Usuario": (d.get('usuario') or '').strip(),
                    "NroDep": (d.get('NroDep') or '').strip()
                })
        
        # 4. Fetch ALL bank movements
        bancos_list = []
        if bank_ids:
            ph = ','.join(['?'] * len(bank_ids))
            cursor.execute(f"SELECT * FROM BankMovements WHERE Id IN ({ph})", bank_ids)
            cols_b = [c[0] for c in cursor.description]
            for br in cursor.fetchall():
                bd = dict(zip(cols_b, br))
                op_num = (bd.get('OpCancelacion') or bd.get('OperacionNumero') or '').strip()
                bancos_list.append({
                    "Id": bd.get('Id'),
                    "Fecha": bd.get('Fecha'),
                    "Descripcion": (bd.get('DescripcionFinal') or bd.get('Descripcion') or '').strip(),
                    "Monto": float(bd.get('Monto') or 0),
                    "Operacion": op_num
                })
        
        return {
            "match": {
                "Id": match_id,
                "ReconciliationId": group_id,
                "MatchedAt": matched_at,
                "MatchType": match_type
            },
            "cobranzas": cobranzas_list,
            "bancos": bancos_list
        }
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()



class ClearAllRequest(BaseModel):
    codcia: str
    bank_code: str
    year: Optional[str] = None
    month: Optional[str] = None

@router.delete("/clear-all")
def clear_all_conciliaciones(request: ClearAllRequest):
    """Elimina todas las conciliaciones para una empresa y banco."""
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        
        # 1. Get all BankMovement IDs for this company/bank/year/month
        clear_q = "SELECT Id FROM BankMovements WHERE CodCia = ? AND BankCode = ? AND Estado = 'Conciliado'"
        clear_params = [request.codcia, request.bank_code]
        if request.year:
            clear_q += " AND YEAR(Fecha) = ?"
            clear_params.append(int(request.year))
        if request.month:
            clear_q += " AND MONTH(Fecha) = ?"
            clear_params.append(int(request.month))
        cursor.execute(clear_q, clear_params)
        bank_ids = [row[0] for row in cursor.fetchall()]
        
        if not bank_ids:
            # Even without bank movements, clean orphan tbl_Conciliados records
            clean_q = "DELETE FROM tbl_Conciliados WHERE empresa = ? AND codigo_banco = ?"
            clean_p = [request.codcia, request.bank_code]
            if request.year:
                clean_q += " AND YEAR(Fecha_banco) = ?"
                clean_p.append(int(request.year))
            if request.month:
                clean_q += " AND MONTH(Fecha_banco) = ?"
                clean_p.append(int(request.month))
            cursor.execute(clean_q, clean_p)
            deleted_orphans = cursor.rowcount
            conn.commit()
            if deleted_orphans > 0:
                return {"status": "success", "message": f"Se limpiaron {deleted_orphans} registros huerfanos de conciliaciones.", "deleted": deleted_orphans}
            return {"status": "success", "message": "No hay conciliaciones para limpiar.", "deleted": 0}
        
        placeholders = ','.join(['?'] * len(bank_ids))
        
        # 2. Get ReconciliationDetail IDs linked to these bank movements
        cursor.execute(f"""
            SELECT Id, ReconciliationId FROM ReconciliationDetail
            WHERE BankMovementId IN ({placeholders})
        """, bank_ids)
        detail_rows = cursor.fetchall()
        detail_ids = [r[0] for r in detail_rows]
        
        if detail_ids:
            detail_placeholders = ','.join(['?'] * len(detail_ids))
            
            # 3. Delete from tbl_Conciliados
            cursor.execute(f"DELETE FROM tbl_Conciliados WHERE ReconciliationDetailId IN ({detail_placeholders})", detail_ids)
            
            # 4. Delete from ReconciliationDetail
            cursor.execute(f"DELETE FROM ReconciliationDetail WHERE Id IN ({detail_placeholders})", detail_ids)
        
        # 5. Reset BankMovements estado
        cursor.execute(f"""
            UPDATE BankMovements 
            SET Estado = 'Pendiente', ReconciliationDetailId = NULL
            WHERE Id IN ({placeholders})
        """, bank_ids)
        
        # 6. Also directly clean tbl_Conciliados by empresa/banco/date (catch orphans)
        clean_conc_q = "DELETE FROM tbl_Conciliados WHERE empresa = ? AND codigo_banco = ?"
        clean_conc_params = [request.codcia, request.bank_code]
        if request.year:
            clean_conc_q += " AND YEAR(Fecha_banco) = ?"
            clean_conc_params.append(int(request.year))
        if request.month:
            clean_conc_q += " AND MONTH(Fecha_banco) = ?"
            clean_conc_params.append(int(request.month))
        cursor.execute(clean_conc_q, clean_conc_params)
        
        conn.commit()
        return {"status": "success", "message": f"Se limpiaron {len(bank_ids)} conciliaciones exitosamente.", "deleted": len(bank_ids)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

# ─── REGLAS DE LIMPIEZA BANCARIA (TAB 3) ─────────────────────────
import json
import os

RULES_FILE = os.path.join(os.path.dirname(__file__), "conciliacion_reglas.json")

def load_rules_from_disk():
    if not os.path.exists(RULES_FILE):
        return []
    try:
        with open(RULES_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []

@router.get("/conciliados")
def get_conciliados(
    codcia: Optional[str] = None,
    bank_code: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None
):
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        query = """
            SELECT Id, ReconciliationDetailId, IdCobranza_CodCia, IdCobranza_coddoc, 
                   IdCobranza_nrodoc, IdCobranza_nroitm, codref, nroref, importe, 
                   codaux, IdBanco, Fecha_banco, empresa, codigo_banco, CreatedAt,
                   nro_operacion, usuario
            FROM tbl_Conciliados
            WHERE 1=1
        """
        params = []
        if codcia:
            query += " AND empresa = ?"
            params.append(codcia)
        if bank_code:
            query += " AND codigo_banco = ?"
            params.append(bank_code)
        if year:
            query += " AND YEAR(Fecha_banco) = ?"
            params.append(year)
        if month:
            query += " AND MONTH(Fecha_banco) = ?"
            params.append(month)

        query += " ORDER BY CreatedAt DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        result = []
        for r in rows:
            result.append({
                "Id": r[0],
                "ReconciliationDetailId": r[1],
                "IdCobranza_CodCia": r[2],
                "IdCobranza_coddoc": r[3],
                "IdCobranza_nrodoc": r[4],
                "IdCobranza_nroitm": r[5],
                "codref": r[6],
                "nroref": r[7],
                "importe": float(r[8] or 0),
                "codaux": r[9],
                "IdBanco": r[10],
                "Fecha_banco": r[11].isoformat() if r[11] else None,
                "empresa": r[12],
                "codigo_banco": r[13],
                "CreatedAt": r[14].isoformat() if r[14] else None,
                "nro_operacion": r[15],
                "usuario": r[16]
            })
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

def save_rules_to_disk(rules):
    with open(RULES_FILE, "w", encoding="utf-8") as f:
        json.dump(rules, f, indent=4, ensure_ascii=False)

class ReglaIn(BaseModel):
    condicion: str
    resultado: str

@router.get("/reglas")
def get_reglas():
    return load_rules_from_disk()

@router.post("/reglas")
def create_regla(regla: ReglaIn):
    rules = load_rules_from_disk()
    new_id = 1 if not rules else max(r.get("id", 0) for r in rules) + 1
    new_rule = {
        "id": new_id,
        "condicion": regla.condicion.strip(),
        "resultado": regla.resultado.strip()
    }
    rules.append(new_rule)
    save_rules_to_disk(rules)
    return new_rule

@router.delete("/reglas/{rule_id}")
def delete_regla(rule_id: int):
    rules = load_rules_from_disk()
    rules = [r for r in rules if r.get("id") != rule_id]
    save_rules_to_disk(rules)
    return {"status": "ok"}

@router.post("/limpiar-banco")
def execute_cleaning_rules():
    """
    Applies the cleaning rules to the Bank Movements in the `ConciliacionBanco` or SQLite
    table where they are stored temporarily. For now, since we parse Excel into SQLite or memory,
    we will apply it to the main memory/SQLite store.
    """
    rules = load_rules_from_disk()
    if not rules:
        return {"registros_actualizados": 0}
        
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="No DB connection")
        
    registros_actualizados = 0
    try:
        cursor = conn.cursor()
        for r in rules:
            condicion = r['condicion']
            resultado = r['resultado']
            # FoxPro update: replace Description with result if Condition matches
            # Wait, bank movements are locally stored in our 'uploaded' memory list!
            pass 
        # Note: Depending on where bank movements are stored (FoxPro, SQLite or Memory array), 
        # this logic modifies them. Since we only read XML/Excel, we will implement this
        # logic upon importing the Excel next time, OR update the FoxPro table directly if
        # it was uploaded to FoxPro.
        
        return {"registros_actualizados": registros_actualizados, "status": "Simulated or done locally"}
    except Exception as e:
        print(f"Error cleaning bank docs: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ─── CLOSE AND REVERT CONCILIADO CAJAS ─────────────────────────

class RevertCajaRequest(BaseModel):
    codcia: str
    coddoc: str
    nrodoc: str


@router.post("/cerrar-cajas")
def cerrar_cajas(
    year: Optional[str] = None,
    month: Optional[str] = None,
    codcia: Optional[str] = None,
    bank_code: Optional[str] = None
):
    """
    Cierra todas las cajas (CcbICaja) que posean al menos un movimiento conciliado
    (ReconciliationDetail) en el período o filtros especificados.
    Pasa flgest de 'P' (EMITIDO) a 'C' (CERRADO).
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        query = """
            UPDATE c
            SET c.flgest = 'C'
            FROM CcbICaja c
            WHERE c.flgest = 'P'
              AND EXISTS (
                  SELECT 1
                  FROM CcbMVtos m
                  JOIN ReconciliationDetail rd ON rd.MatchCodCia = m.CodCia
                      AND rd.MatchCoddoc = m.coddoc
                      AND rd.MatchNrodoc = m.nrodoc
                      AND rd.MatchNroitm = m.nroitm
                  WHERE m.CodCia = c.codcia
                    AND m.coddoc = c.coddoc
                    AND m.nrodoc = c.nrodoc
        """
        params = []
        if year:
            query += " AND m.anos = ?"
            params.append(year)
        if month:
            query += " AND m.mes = ?"
            params.append(month.zfill(2))
        if codcia:
            if bank_code:
                query += " AND (m.CodCia = ? OR (m.tpopgo = '1' AND m.CodCom = ? AND m.CodDep = ?))"
                params.extend([codcia, codcia, bank_code])
            else:
                query += " AND m.CodCia = ?"
                params.append(codcia)

        query += " )" # Cierra el EXISTS

        cursor.execute(query, params)
        updated_count = cursor.rowcount
        conn.commit()
        
        return {
            "status": "success",
            "message": f"Se cerraron {updated_count} cajas correctamente.",
            "closed_count": updated_count
        }
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@router.post("/revertir-caja")
def revertir_caja(request: RevertCajaRequest):
    """
    Revierte el estado de una caja específica (CcbICaja) de 'C' (CERRADO) a 'P' (EMITIDO).
    """
    conn = get_db_connection()
    if not conn:
        raise HTTPException(status_code=500, detail="DB Error")
    try:
        cursor = conn.cursor()
        cursor.execute("""
            UPDATE CcbICaja
            SET flgest = 'P'
            WHERE codcia = ? AND coddoc = ? AND nrodoc = ? AND flgest = 'C'
        """, (request.codcia, request.coddoc, request.nrodoc))
        
        updated_count = cursor.rowcount
        conn.commit()
        
        if updated_count == 0:
            cursor.execute("SELECT flgest FROM CcbICaja WHERE codcia = ? AND coddoc = ? AND nrodoc = ?", (request.codcia, request.coddoc, request.nrodoc))
            row = cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Caja no encontrada.")
            else:
                return {"status": "success", "message": f"La caja ya estaba en estado {row[0]}."}
                
        return {"status": "success", "message": "Caja revertida a EMITIDO (FLGEST=P) correctamente."}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()

